"""Generic file extractor for various file types."""

from __future__ import annotations

import hashlib
import time
from datetime import datetime
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from ingestion.extractors.base import BaseExtractor
from ingestion.metadata import ExtractedContent, SourceMetadata


class FileExtractor(BaseExtractor):
    """Extract content from various file types."""

    def __init__(self):
        """Initialize file extractor."""
        self.supported_extensions = {
            ".txt": self._extract_text,
            ".md": self._extract_text,
            ".docx": self._extract_docx,
            ".xlsx": self._extract_xlsx,
            ".csv": self._extract_csv,
            ".json": self._extract_json,
            # Images, audio, video are handled by processors
            ".png": self._extract_image_metadata,
            ".jpg": self._extract_image_metadata,
            ".jpeg": self._extract_image_metadata,
            ".gif": self._extract_image_metadata,
            ".mp4": self._extract_video_metadata,
            ".mp3": self._extract_audio_metadata,
            ".wav": self._extract_audio_metadata,
        }

    def can_handle(self, source: str) -> bool:
        """Check if source is a file path."""
        path = Path(source)
        return path.exists() and path.is_file()

    def extract(self, source: str, **kwargs) -> ExtractedContent:
        """Extract content from file."""
        start_time = time.time()
        path = Path(source)

        source_id = hashlib.sha1(str(path.absolute()).encode()).hexdigest()
        ext = path.suffix.lower()

        # Determine source type
        if ext in [".png", ".jpg", ".jpeg", ".gif"]:
            source_type = "image"
        elif ext in [".mp4", ".avi", ".mov"]:
            source_type = "video"
        elif ext in [".mp3", ".wav", ".m4a"]:
            source_type = "audio"
        else:
            source_type = "file"

        metadata = SourceMetadata(
            source_type=source_type,
            source_url=str(path.absolute()),
            source_id=source_id,
            ingested_at=datetime.utcnow(),
            original_format=ext.lstrip("."),
            mime_type=self._detect_mime_type(path),
            extraction_method="file_extraction",
        )

        # Extract based on extension
        extractor_func = self.supported_extensions.get(ext)
        if extractor_func:
            text = extractor_func(path)
        else:
            # Fallback: try to read as text
            try:
                text = path.read_text(encoding="utf-8", errors="replace")
            except Exception:
                text = f"[File: {path.name} - Content extraction not supported for this file type]"

        duration = time.time() - start_time
        metadata.processing_duration_seconds = duration
        metadata.processing_steps.append("file_extraction")

        return ExtractedContent(
            text=text,
            metadata=metadata,
            raw_content_path=path,
        )

    def _extract_text(self, path: Path) -> str:
        """Extract text from plain text file."""
        return path.read_text(encoding="utf-8", errors="replace")

    def _extract_docx(self, path: Path) -> str:
        """Extract text from DOCX file."""
        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)

    def _extract_xlsx(self, path: Path) -> str:
        """Extract text from Excel file."""
        wb = load_workbook(path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}\n")
            rows = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell for cell in row):
                    cells = [str(cell or "") for cell in row]
                    rows.append(" | ".join(cells))
            parts.append("\n".join(rows))
        return "\n\n".join(parts)

    def _extract_csv(self, path: Path) -> str:
        """Extract text from CSV file."""
        import csv

        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = [" | ".join(row) for row in reader]
        return "\n".join(rows)

    def _extract_json(self, path: Path) -> str:
        """Extract text from JSON file (formatted)."""
        import json

        data = json.loads(path.read_text(encoding="utf-8"))
        return json.dumps(data, indent=2)

    def _extract_image_metadata(self, path: Path) -> str:
        """Extract metadata from image (actual processing done by processor)."""
        from PIL import Image

        try:
            img = Image.open(path)
            return f"[Image: {path.name}, Size: {img.size}, Format: {img.format}]"
        except Exception:
            return f"[Image: {path.name}]"

    def _extract_video_metadata(self, path: Path) -> str:
        """Extract metadata from video (actual processing done by processor)."""
        try:
            import ffmpeg

            probe = ffmpeg.probe(str(path))
            video_info = probe.get("streams", [{}])[0]
            duration = probe.get("format", {}).get("duration", "unknown")
            return f"[Video: {path.name}, Duration: {duration}s, Codec: {video_info.get('codec_name', 'unknown')}]"
        except Exception:
            return f"[Video: {path.name}]"

    def _extract_audio_metadata(self, path: Path) -> str:
        """Extract metadata from audio (actual processing done by processor)."""
        return f"[Audio: {path.name}]"
